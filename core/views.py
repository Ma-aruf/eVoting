from io import BytesIO
import logging
import sys

from django.contrib.auth import get_user_model
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import viewsets, status
from rest_framework.exceptions import ParseError
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .authentication import VoterAuthentication
from .models import Election, Position, Candidate, Vote, Student
from .permissions import IsStaffOrSuperUser, IsActivatorOrSuperUser, IsStaffOrSuperUserOrReadOnlyActivator, IsSuperUser
from .serializers import (
    StudentSerializer,
    BulkStudentUploadSerializer,
    ElectionSerializer,
    PositionSerializer,
    CandidateSerializer,
    MultiVoteSerializer,
    UserSerializer,
)
from .utils import generate_voter_hmac

User = get_user_model()


class UserViewSet(viewsets.ModelViewSet):
    """
    Superuser-only: manage admin users (staff, activator, superuser).
    """
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    permission_classes = [IsSuperUser]

    def get_queryset(self):
        return User.objects.all().order_by('-date_joined')


class ElectionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Public read-only access to elections.
    Voters need to see active elections without JWT auth.
    """
    queryset = Election.objects.all()
    serializer_class = ElectionSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        # Optional: allow filtering by is_active
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            return Election.objects.filter(is_active=is_active.lower() == 'true')
        return Election.objects.all()


class StudentViewSet(viewsets.ModelViewSet):
    """
    Staff or superuser can manage students (CRUD).
    Activator is intentionally excluded from create/update/delete and instead
    uses StudentActivationView to only toggle activation.
    """
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsStaffOrSuperUserOrReadOnlyActivator]

    def get_queryset(self):
        election_id = self.request.query_params.get("election_id")
        if election_id:
            return Student.objects.filter(election_id=election_id)
        return Student.objects.all()

    def perform_create(self, serializer):
        """Ensure election is set when creating a student."""
        election_id = self.request.data.get('election_id')
        if not election_id:
            raise ParseError("election_id is required for student creation.")
        
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            raise ParseError("Invalid election_id provided.")
        
        serializer.save(election=election)

    def destroy(self, request, *args, **kwargs):
        student = self.get_object()
        if student.has_voted:
            return Response(
                {"detail": "Cannot delete a student who has already voted."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)


class BulkStudentUploadView(APIView):
    """
    Allow staff/superuser to upload an Excel file to create students in bulk.
    Expected columns (case-insensitive): student_id, full_name, class_name.
    """

    permission_classes = [IsStaffOrSuperUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        serializer = BulkStudentUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = serializer.validated_data["file"]

        # Accept optional election_id from request; if not provided, reject
        election_id = request.data.get("election_id")
        if not election_id:
            return Response(
                {"detail": "election_id is required for bulk upload."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            return Response(
                {"detail": "Invalid election_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(filename=BytesIO(upload.read()), read_only=True)
            ws = wb.active
        except Exception:
            return Response(
                {"detail": "Could not read Excel file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Map headers to indices
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header_map = {str(h or "").strip().lower(): idx for idx, h in enumerate(header_row)}
        required = ["student_id", "full_name", "class_name"]
        missing = [h for h in required if h not in header_map]
        if missing:
            return Response(
                {"detail": f"Missing columns: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows_to_create = []
        file_ids = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Force everything to string right away (handles int/float/None nicely)
            get_str = lambda idx: str(row[idx]).strip() if row[idx] is not None else ""

            student_id = get_str(header_map["student_id"])
            full_name = get_str(header_map["full_name"])
            class_name = get_str(header_map["class_name"])

            if not student_id or not full_name or not class_name:
                continue  # skip incomplete rows

            if student_id in file_ids:
                continue  # skip duplicates in the same file
            file_ids.add(student_id)

            rows_to_create.append(
                Student(
                    student_id=student_id,
                    full_name=full_name,
                    class_name=class_name,
                    election=election,
                )
            )

        if not rows_to_create:
            return Response(
                {"detail": "No valid rows found to import."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existing_ids = set(
            Student.objects.filter(
                student_id__in=[s.student_id for s in rows_to_create],
                election_id=election_id  # Only check within this election
            )
            .values_list("student_id", flat=True)
        )

        rows_to_create = [s for s in rows_to_create if s.student_id not in existing_ids]

        if not rows_to_create:
            return Response(
                {"detail": "All provided students already exist."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            Student.objects.bulk_create(rows_to_create, ignore_conflicts=True)
            return Response(
                {
                    "detail": "Students imported successfully.",
                    "created": len(rows_to_create),
                    "skipped_existing": len(existing_ids),
                    "election": election.name,
                },
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            return Response(
                {"detail": f"Bulk import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )


class PositionViewSet(viewsets.ModelViewSet):
    """
    Read positions publicly, but only staff/superuser can update/delete.
    Expects `?election_id=` as a query parameter for listing.
    """
    serializer_class = PositionSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsStaffOrSuperUser()]

    def get_queryset(self):
        election_id = self.request.query_params.get("election_id")
        if election_id:
            return Position.objects.filter(election_id=election_id)
        return Position.objects.all()


class PositionCreateView(APIView):
    """
    Staff or superuser can create positions for an election.
    """
    permission_classes = [IsStaffOrSuperUser]

    def post(self, request):
        serializer = PositionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        position = serializer.save()
        return Response(PositionSerializer(position).data, status=status.HTTP_201_CREATED)

    def put(self, request, pk):
        position = get_object_or_404(Position, pk=pk)
        serializer = PositionSerializer(
            position,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        position = serializer.save()
        return Response(
            PositionSerializer(position).data,
            status=status.HTTP_200_OK
        )

    def patch(self, request, pk):
        return self.put(request, pk)

    def delete(self, request, pk):
        position = get_object_or_404(Position, pk=pk)
        position.delete()
        return Response(
            {"detail": "Position deleted successfully"},
            status=status.HTTP_204_NO_CONTENT
        )


class CandidateViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Public, read-only list of candidates for a given position.
    Expects `?position_id=` as a query parameter.
    """
    serializer_class = CandidateSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        position_id = self.request.query_params.get("position_id")
        if position_id:
            return Candidate.objects.filter(position_id=position_id).order_by('ballot_number')
        return Candidate.objects.none()


class CandidateCreateView(APIView):
    """
    Staff or superuser can register candidates for positions.
    """
    permission_classes = [IsStaffOrSuperUser]

    # CREATE
    def post(self, request):
        serializer = CandidateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        candidate = serializer.save()
        return Response(CandidateSerializer(candidate).data, status=status.HTTP_201_CREATED)

    # EDIT
    def put(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        serializer = CandidateSerializer(
            candidate,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        candidate = serializer.save()
        return Response(
            CandidateSerializer(candidate).data,
            status=status.HTTP_200_OK
        )

    def patch(self, request, pk):
        return self.put(request, pk)

    # DELETE
    def delete(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        candidate.delete()
        return Response(
            {"detail": "Candidate deleted successfully"},
            status=status.HTTP_204_NO_CONTENT
        )


class ElectionManageView(APIView):
    """
    Staff or superuser can start/stop elections by toggling `is_active`.
    """

    permission_classes = [IsStaffOrSuperUser]
    
    security_logger = logging.getLogger('security')

    def get(self, request):
        # Return all elections (active and inactive)
        elections = Election.objects.all().order_by('-year', '-start_time')
        serializer = ElectionSerializer(elections, many=True)
        return Response(serializer.data)

    def patch(self, request):
        """
        Accepts JSON: { "election_id": 1, "is_active": true }
        Multiple elections can be active simultaneously.
        """
        print("Inside post")
        election_id = request.data.get("election_id")
        is_active = request.data.get("is_active")
        
        # Get client IP and user for logging
        client_ip = request.META.get('REMOTE_ADDR')
        user = request.user

        if election_id is None:
            return Response(
                {"detail": "election_id required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if is_active is None:
            return Response(
                {"detail": "is_active required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            return Response(
                {"detail": "Election not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        with transaction.atomic():
            # Allow multiple elections to be active simultaneously
            # Students are scoped by election_id, so no vote mixing occurs
            election.is_active = bool(is_active)
            election.save(update_fields=["is_active"])
            
            # Log election status change
            action = "STARTED" if bool(is_active) else "STOPPED"
            self.security_logger.info(
                f"ELECTION_{action}: election_id={election_id}, election_name={election.name}, "
                f"user={user.username if user else 'unknown'}, ip={client_ip}"
            )

        return Response(
            {"detail": "Election status updated.", "id": election.pk, "is_active": election.is_active},
            status=status.HTTP_200_OK,
        )


class MultiVoteView(APIView):
    """
    Students authenticate via headers using `VoterAuthentication`.
    View enforces activation, single-vote, and transactional locking.
    """
    authentication_classes = [VoterAuthentication]
    permission_classes = [IsAuthenticated]
    
    security_logger = logging.getLogger('security')
    
    def post(self, request):
        # Apply rate limiting only in production
        from django.conf import settings
        if getattr(settings, 'RATE_LIMITING_ENABLED', False):
            from django_ratelimit.decorators import ratelimit
            from django.utils.decorators import method_decorator
            
            @method_decorator(ratelimit(key='ip', rate='10/m', method='POST'))
            def rate_limited_post(self, request):
                return self._actual_post(request)
            return rate_limited_post(self, request)
        else:
            return self._actual_post(request)
    
    def _actual_post(self, request):
        serializer = MultiVoteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        # `request.user` is StudentUser from VoterAuthentication
        student_user = getattr(request.user, "student", None)
        token = getattr(request, "auth", None)
        if student_user is None or token is None:
            raise ParseError("Student authentication required via headers.")

        # Get client IP for logging
        client_ip = request.META.get('REMOTE_ADDR')
        
        # Log vote attempt
        self.security_logger.info(
            f"VOTE_ATTEMPT: student_id={student_user.student_id if student_user else 'unknown'}, "
            f"ip={client_ip}, election_ids={[v['election'] for v in data['votes']]}"
        )

        try:
            with transaction.atomic():
                # Lock fresh student row to avoid races
                student = Student.objects.select_for_update().get(
                    pk=student_user.pk
                )

                now = timezone.now()

                if not getattr(student, "is_active", False):
                    self.security_logger.warning(
                        f"VOTE_DENIED_INACTIVE: student_id={student.student_id}, ip={client_ip}"
                    )
                    return Response(
                        {"detail": "Student is not activated to vote."},
                        status=status.HTTP_403_FORBIDDEN,
                    )

                if getattr(student, "has_voted", False):
                    self.security_logger.warning(
                        f"VOTE_DENIED_ALREADY_VOTED: student_id={student.student_id}, ip={client_ip}"
                    )
                    return Response(
                        {"detail": "Student has already voted."},
                        status=status.HTTP_403_FORBIDDEN,
                    )

                votes_to_create = []
                election_cache = {}
                position_cache = {}
                candidate_cache = {}
                for vote_data in data["votes"]:
                    election_id = vote_data["election"]
                    position_id = vote_data["position"]
                    candidate_id = vote_data["candidate"]

                    # Load and validate election (must be active and within window)
                    election = election_cache.get(election_id)
                    if election is None:
                        election = (
                            Election.objects.filter(pk=election_id, is_active=True)
                            .select_for_update()
                            .first()
                        )
                        election_cache[election_id] = election
                    if election is None:
                        return Response(
                            {"detail": "Election is not active or does not exist."},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                    if election.start_time > now or election.end_time < now:
                        return Response(
                            {"detail": "Election is not within the voting window."},
                            status=status.HTTP_403_FORBIDDEN,
                        )

                    # Validate position belongs to election
                    position = position_cache.get(position_id)
                    if position is None:
                        position = (
                            Position.objects.filter(
                                pk=position_id, election_id=election.pk
                            )
                            .select_for_update()
                            .first()
                        )
                        position_cache[position_id] = position
                    if position is None:
                        return Response(
                            {"detail": "Position does not belong to election."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    # Validate candidate belongs to position
                    candidate = candidate_cache.get(candidate_id)
                    if candidate is None:
                        candidate = (
                            Candidate.objects.filter(
                                pk=candidate_id, position_id=position.pk
                            )
                            .select_for_update()
                            .first()
                        )
                        candidate_cache[candidate_id] = candidate
                    if candidate is None:
                        return Response(
                            {"detail": "Candidate does not belong to position."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    # Ensure no existing vote for that position by this voter token
                    if Vote.objects.filter(
                            voter_hash=token, position_id=position_id
                    ).exists():
                        return Response(
                            {"detail": "Duplicate vote detected for a position."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    votes_to_create.append(
                        Vote(
                            voter_hash=token,
                            election_id=election_id,
                            position_id=position_id,
                            candidate_id=candidate_id,
                        )
                    )

                Vote.objects.bulk_create(votes_to_create)

                # Mark student as voted and deactivate
                student.has_voted = True
                student.is_active = False
                student.save(update_fields=["has_voted", "is_active"])
                
                # Log successful vote
                self.security_logger.info(
                    f"VOTE_SUCCESS: student_id={student.student_id}, ip={client_ip}, "
                    f"votes_count={len(votes_to_create)}, election_ids={[v.election_id for v in votes_to_create]}"
                )

        except Student.DoesNotExist:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {"detail": "All votes submitted successfully."},
            status=status.HTTP_201_CREATED,
        )


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            "username": user.username,
            "role": user.role,
        })


class StudentActivationView(APIView):
    """
    Toggle `is_active` on a Student. Only activator or superuser may call.
    Accepts JSON: { "student_id": "S12345", "election_id": 1, "is_active": true }
    """
    permission_classes = [IsActivatorOrSuperUser]
    
    security_logger = logging.getLogger('security')
    
    def post(self, request):
        # Apply rate limiting only in production
        from django.conf import settings
        if getattr(settings, 'RATE_LIMITING_ENABLED', False):
            from django_ratelimit.decorators import ratelimit
            from django.utils.decorators import method_decorator
            
            @method_decorator(ratelimit(key='ip', rate='11/m', method='POST'))
            def rate_limited_post(self, request):
                return self._actual_post(request)
            return rate_limited_post(self, request)
        else:
            return self._actual_post(request)
    
    def _actual_post(self, request):
        print("INSIDE ACTIVATION VIEW - POST CALLED")  # ← add this
        print(request.path, request.method)
        
        # Get client IP and user for logging
        client_ip = request.META.get('REMOTE_ADDR')
        user = request.user
        
        student_id = request.data.get("student_id")
        election_id = request.data.get("election_id")
        
        # Log activation attempt
        self.security_logger.info(
            f"ACTIVATION_ATTEMPT: student_id={student_id}, election_id={election_id}, "
            f"user={user.username if user else 'unknown'}, ip={client_ip}"
        )
        
        if not student_id:
            return Response(
                {"detail": "student_id required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
            
        if not election_id:
            return Response(
                {"detail": "election_id required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_active = request.data.get("is_active")
        if is_active is None:
            return Response(
                {"detail": "is_active required."},
                status=status.HTTP_400_BAD_REQUEST,
            )



        try:
            # Get the specific election
            election = Election.objects.get(id=election_id)
        except Election.DoesNotExist:
            return Response(
                {"detail": "Election not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            # Find student by both student_id and election_id
            student = Student.objects.get(student_id=student_id, election_id=election_id)
        except Student.DoesNotExist:
            return Response(
                {"detail": "Student not found in this election."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Check if student has already voted (cannot be activated if voted)
        if student.has_voted:
            self.security_logger.warning(
                f"ACTIVATION_DENIED_VOTED: student_id={student_id}, election_id={election_id}, "
                f"user={user.username if user else 'unknown'}, ip={client_ip}"
            )
            return Response(
                {"detail": "Student has already voted and cannot be re-activated."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check current status for better message
        current_status = student.is_active
        new_status = bool(is_active)

        if current_status == new_status:
            status_text = "active" if current_status else "inactive"
            return Response(
                {"detail": f"Student is already {status_text}."},
                status=status.HTTP_200_OK,
            )

        # Only toggle the is_active flag
        student.is_active = new_status
        student.save(update_fields=["is_active"])
        
        # Log successful activation/deactivation
        action = "ACTIVATED" if new_status else "DEACTIVATED"
        self.security_logger.info(
            f"STUDENT_{action}: student_id={student_id}, election_id={election_id}, "
            f"user={user.username if user else 'unknown'}, ip={client_ip}"
        )

        status_text = "activated" if new_status else "deactivated"
        return Response(
            {"detail": f"Student {status_text} successfully."},
            status=status.HTTP_200_OK,
        )


class ElectionCreateView(APIView):
    """
    Staff or superuser can create a new election.
    """
    permission_classes = [IsStaffOrSuperUser]

    def post(self, request):
        serializer = ElectionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        election = serializer.save()
        return Response(ElectionSerializer(election).data, status=status.HTTP_201_CREATED)


class StudentVoterLoginView(APIView):
    """
    Generate HMAC token for active students who haven't voted yet.
    """
    permission_classes = [AllowAny]
    
    security_logger = logging.getLogger('security')
    
    def post(self, request):
        # Apply rate limiting only in production
        from django.conf import settings
        
        if getattr(settings, 'RATE_LIMITING_ENABLED', False):
            from django_ratelimit.decorators import ratelimit
            from django.utils.decorators import method_decorator
            
            @method_decorator(ratelimit(key='ip', rate='5/m', method='POST'))
            def rate_limited_post(self, request):
                return self._actual_post(request)
            return rate_limited_post(self, request)
        else:
            return self._actual_post(request)
    
    def _actual_post(self, request):
        student_id = request.data.get("student_id")
        
        # Get client IP for logging
        client_ip = request.META.get('REMOTE_ADDR')
        
        # Log login attempt
        self.security_logger.info(
            f"LOGIN_ATTEMPT: student_id={student_id}, ip={client_ip}"
        )

        if not student_id:
            return Response(
                {"detail": "student_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()

        # Find all active elections within voting window
        active_elections = Election.objects.filter(
            is_active=True,
            start_time__lte=now,
            end_time__gte=now
        )

        if not active_elections.exists():
            return Response(
                {"detail": "No active election at this time."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Find the student who is ACTIVATED (is_active=True) in any of the active elections
        # Student identity = student_id + election_id (composite)
        # A student can only be activated in one election at a time
        try:
            student = Student.objects.get(
                student_id=student_id,
                election__in=active_elections,
                is_active=True,  # Must be activated to vote
                has_voted=False  # Must not have voted yet
            )
            active_election = student.election
        except Student.DoesNotExist:
            # Check if student exists but is not activated or has voted
            existing_student = Student.objects.filter(
                student_id=student_id,
                election__in=active_elections
            ).first()
            
            if existing_student:
                if existing_student.has_voted:
                    self.security_logger.warning(
                        f"LOGIN_DENIED_VOTED: student_id={student_id}, ip={client_ip}"
                    )
                    return Response(
                        {"detail": "Student has already voted."},
                        status=status.HTTP_409_CONFLICT,
                    )
                else:
                    self.security_logger.warning(
                        f"LOGIN_DENIED_INACTIVE: student_id={student_id}, ip={client_ip}"
                    )
                    return Response(
                        {"detail": "Student is not activated to vote."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            self.security_logger.warning(
                f"LOGIN_NOT_FOUND: student_id={student_id}, ip={client_ip}"
            )
            return Response(
                {"detail": "Student not found in any active election."},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Student.MultipleObjectsReturned:
            # Edge case: same student_id activated in multiple elections simultaneously
            return Response(
                {"detail": "Student is activated in multiple elections. Please contact administrator."},
                status=status.HTTP_409_CONFLICT,
            )

        # Generate HMAC token (include election_id to scope the token)
        token = generate_voter_hmac(f"{student.student_id}_{active_election.id}")
        
        # Log successful login
        self.security_logger.info(
            f"LOGIN_SUCCESS: student_id={student.student_id}, election_id={active_election.id}, ip={client_ip}"
        )

        return Response({
            "token": token,
            "student": {
                "id": student.id,
                "student_id": student.student_id,
                "full_name": student.full_name,
                "class_name": student.class_name,
            },
            "election": {
                "id": active_election.id,
                "name": active_election.name,
                "year": active_election.year,
            }
        }, status=status.HTTP_200_OK)


class ElectionStatsView(APIView):
    """Get basic election statistics"""
    permission_classes = [IsStaffOrSuperUser]

    def get(self, request, election_id):
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            return Response(
                {"detail": "Election not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        total_voters = Student.objects.filter(election=election).count()
        voters_voted = Student.objects.filter(election=election, has_voted=True).count()

        return Response({
            "election_id": election.id,
            "election_name": election.name,
            "total_voters": total_voters,
            "voters_voted": voters_voted,
            "turnout_percentage": round((voters_voted / total_voters * 100), 2) if total_voters > 0 else 0.0
        })


class PositionStatsView(APIView):
    """Get statistics for a specific position including skipped votes"""
    permission_classes = [IsStaffOrSuperUser]

    def get(self, request):
        position_id = request.query_params.get("position_id")
        if not position_id:
            return Response(
                {"detail": "position_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            position = Position.objects.get(pk=position_id)
        except Position.DoesNotExist:
            return Response(
                {"detail": "Position not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        election = position.election

        # Unique voters who cast any vote in this election
        unique_voters = Vote.objects.filter(election=election).values('voter_hash').distinct().count()

        # Votes actually cast for this position
        position_votes = Vote.objects.filter(position=position).count()

        skipped = max(0, unique_voters - position_votes)

        return Response({
            "position_id": position.id,
            "position_name": position.name,
            "election": election.name,
            "unique_voters_in_election": unique_voters,
            "votes_for_this_position": position_votes,
            "skipped_votes": skipped,
            "skip_percentage": round((skipped / unique_voters * 100), 2) if unique_voters > 0 else 0.0
        })


class ElectionResultsView(APIView):
    """Get comprehensive results for an entire election"""
    permission_classes = [IsStaffOrSuperUser]

    def get(self, request, election_id):
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            return Response(
                {"detail": "Election not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # All positions in display order
        positions = Position.objects.filter(election=election).order_by('display_order')

        # Total unique voters in this election (across all positions)
        unique_voters = Vote.objects.filter(election=election).values('voter_hash').distinct().count()

        results = []

        for position in positions:
            candidates = Candidate.objects.filter(position=position).order_by('ballot_number')

            candidate_results = []
            total_valid_votes_this_position = 0

            for candidate in candidates:
                vote_count = Vote.objects.filter(
                    candidate=candidate,
                    position=position
                ).count()

                candidate_results.append({
                    "id": candidate.id,
                    "student_id": candidate.student.student_id,
                    "candidate_name": candidate.student.full_name,
                    "photo_url": candidate.photo_url or "",
                    "vote_count": vote_count,
                })

                total_valid_votes_this_position += vote_count

            skipped = max(0, unique_voters - total_valid_votes_this_position)

            # Add percentages
            for cand in candidate_results:
                cand["percentage"] = (
                    round((cand["vote_count"] / total_valid_votes_this_position * 100), 2)
                    if total_valid_votes_this_position > 0 else 0.0
                )

            # Sort candidates by votes descending
            candidate_results.sort(key=lambda x: x["vote_count"], reverse=True)

            results.append({
                "position_id": position.id,
                "position_name": position.name,
                "display_order": position.display_order,
                "total_valid_votes": total_valid_votes_this_position,
                "skipped_votes": skipped,
                "skip_percentage": round((skipped / unique_voters * 100), 2) if unique_voters > 0 else 0.0,
                "candidates": candidate_results,
            })

        # Overall election stats
        total_students = Student.objects.filter(election=election).count()
        students_who_voted = Student.objects.filter(election=election, has_voted=True).count()

        return Response({
            "election_id": election.id,
            "election_name": election.name,
            "year": election.year,
            "total_students": total_students,
            "students_who_voted": students_who_voted,
            "voter_turnout_percentage": round((students_who_voted / total_students * 100),
                                              2) if total_students > 0 else 0.0,
            "unique_voters_who_cast_at_least_one_vote": unique_voters,
            "positions": results,
        })


class CandidatesForPositionView(APIView):
    permission_classes = [IsStaffOrSuperUser]

    def get(self, request):
        position_id = request.query_params.get("position_id")
        print(position_id)

        if not position_id:
            return Response(
                {"detail": "position_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            candidates = Candidate.objects.filter(position_id=position_id).select_related('student').order_by('ballot_number')
        except ValueError:
            return Response(
                {"detail": "Invalid position_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = []

        for candidate in candidates:
            vote_count = Vote.objects.filter(
                candidate_id=candidate.id,
                position_id=position_id
            ).count()

            candidate_data = {
                "candidate_id": candidate.id,
                "candidate_name": candidate.student.full_name,
                "student_id": candidate.student.student_id,
                "positionid": int(position_id),  # Add position_id to response
                "vote_count": vote_count
            }

            result.append(candidate_data)

        return Response(result, status=status.HTTP_200_OK)


class ImageUploadView(APIView):
    """
    Upload candidate photos.
    - In production (Cloudinary configured): uploads to Cloudinary
    - In development (no Cloudinary): saves to local media folder
    """
    permission_classes = [IsStaffOrSuperUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        from django.conf import settings
        import uuid
        import os

        file = request.FILES.get('image')
        if not file:
            return Response(
                {"detail": "No image file provided."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file type
        allowed_types = ['image/jpeg', 'image/png', 'image/webp']
        if file.content_type not in allowed_types:
            return Response(
                {"detail": "Invalid file type. Allowed: JPEG, PNG, WebP."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file size (max 5MB)
        max_size = 5 * 1024 * 1024
        if file.size > max_size:
            return Response(
                {"detail": "File too large. Maximum size is 5MB."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if Cloudinary is configured (production)
        cloud_name = settings.CLOUDINARY_CLOUD_NAME
        api_key = settings.CLOUDINARY_API_KEY
        api_secret = settings.CLOUDINARY_API_SECRET

        if cloud_name and api_key and api_secret:
            # Production: Upload to Cloudinary
            try:
                import cloudinary
                import cloudinary.uploader

                cloudinary.config(
                    cloud_name=cloud_name,
                    api_key=api_key,
                    api_secret=api_secret
                )

                result = cloudinary.uploader.upload(
                    file,
                    folder="evoting/candidates",
                    transformation=[
                        {"width": 400, "height": 400, "crop": "fill", "gravity": "face"}
                    ]
                )

                return Response({
                    "url": result["secure_url"],
                    "public_id": result["public_id"],
                    "storage": "cloudinary"
                }, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response(
                    {"detail": f"Cloudinary upload failed: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            # Development: Save to local media folder
            try:
                # Ensure media directory exists
                media_path = os.path.join(settings.MEDIA_ROOT, 'candidates')
                os.makedirs(media_path, exist_ok=True)

                # Generate unique filename
                ext = file.name.split('.')[-1] if '.' in file.name else 'jpg'
                filename = f"{uuid.uuid4().hex}.{ext}"
                filepath = os.path.join(media_path, filename)

                # Save file
                with open(filepath, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)

                # Return full URL (include host for frontend to access)
                # Build absolute URL from request
                relative_url = f"{settings.MEDIA_URL}candidates/{filename}"
                url = request.build_absolute_uri(relative_url)

                return Response({
                    "url": url,
                    "filename": filename,
                    "storage": "local"
                }, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response(
                    {"detail": f"Local upload failed: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
